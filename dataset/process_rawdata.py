data_r = open('./dataset/data.txt', mode='rt', encoding='utf-8')
errcode_r = open('./dataset/errorcode.txt', mode='rt', encoding='utf-8')

read_data = data_r.read()
read_errcode = errcode_r.read()

raw_errcode = read_errcode.split('\n')

errcode_map = dict()

for errcode_data in raw_errcode:
    errcode_data_split = errcode_data.split(' ', 1)
    errcode_map[errcode_data_split[0].strip()] = errcode_data_split[1].strip()

# print(errcode_map)

class ErrorData:
    def __init__(self):
        self.situation = ''
        self.cause = ''
        self.description = ''
        self.solution = ''
        self.error_code = ''

    def set_data(self, col, data):
        if col == '현상':
            self.situation += data
        elif col == '원인':
            self.cause += data
        elif col == '조치':
            self.solution += data
        elif col == 'code':
            self.error_code += data
        elif col == 'kor_desc':
            self.description += data
        # self.description = data
        # self.error_code = data

    def set_errcode(self, errcode_list):
        error_code = ''
        description = ''
        for code, desc in errcode_list.items():
            error_code += code + '\n'
            description += desc + '\n'

        self.error_code = error_code
        self.description = description

    def reset(self):
        self.situation = ''
        self.cause = ''
        self.description = ''
        self.solution = ''
        self.error_code = ''

    def print_data(self):
        print('====================================================')
        print('situation:', self.situation)
        print('cause:', self.cause)
        print('solution:', self.solution)
        print('description:', self.description)
        print('error_code:', self.error_code)


def extract_errcode(text):
    #     text_split = text.split(' ')
    fst_idx = -1
    last_idx = -1
    errcode = ''
    fst_idx = text.find('ORA-')
    if (fst_idx == -1):
        return errcode

    errcode = text[fst_idx:fst_idx + 9]
    errcode = errcode.strip()

    last_idx = len(errcode)

    fixed = False
    for idx, c in enumerate(errcode[4:]):
        if not (c >= '0' and c <= '9'):
            last_idx = idx
            True
            break

    errcode = errcode[:last_idx]
    #     print(fixed, 'res:', errcode, 'len:', len(errcode))
    return errcode


def extract_errcode_list(text, errcode_map):
    text_split = text.split('\n')
    errcode_list = dict()
    for line in text_split:
        # 해당 라인에 ORA로 시작하는 에러코드가 있는지 검사
        fst_idx = line.find('ORA-')
        if (fst_idx == -1):
            continue

        # 에러코드부분 추출(최대 9글자)
        errcode = line[fst_idx:fst_idx + 9]
        errcode = errcode.strip()

        # 코드 이후 붙은 문자 제거
        last_idx = len(errcode)
        for idx, c in enumerate(errcode[4:]):
            if not (c >= '0' and c <= '9'):
                last_idx = idx + 4
                break
        errcode = errcode[:last_idx]

        if len(errcode) < 9:
            errcode = errcode[:4] + '0' * (9 - len(errcode)) + errcode[4:9]

        # 에러코드 한글설명 바인딩
        errcode_desc = ''
        if errcode in errcode_map:
            errcode_desc = errcode_map[errcode]

        if errcode != '':
            errcode_list[errcode] = errcode_desc
    #             errcode_list.append({errcode : errcode_desc})
    #     print(fixed, 'res:', errcode, 'len:', len(errcode))
    return errcode_list


import copy

raw_data = read_data
raw_data_split = raw_data.split('\n')

split_data = []
for data in raw_data_split:
    if not data.startswith('-'):  # ------------- 텍스트 제거
        split_data.append(data)

cur_col = None
cur_data = ErrorData()
err_datalist = []
for data in split_data:
    data_split = data.split(':', 1)
    if cur_col is None:
        cur_col = data_split[0].strip()
        # 에러코드 처리
        errcode_list = extract_errcode(data_split[1])
        cur_data.set_data(col=cur_col, data=data_split[1].strip() + '\n')
    elif not (data_split[0].strip().startswith('현상') or
              data_split[0].strip().startswith('조치') or
              data_split[0].strip().startswith('원인')):
        cur_data.set_data(col=cur_col, data=data + '\n')
        # 에러코드 처리
    #         errcode = extract_errcode(data_split[0])
    #         cur_data.set_data(col='code', data=data+'\n')
    #         if errcode in errcode_map:
    #             cur_data.set_data(col='kor_desc', data=data+'\n')
    else:
        if not cur_col.startswith(data_split[0].strip()):  # col 바뀔때(현상, 조치, 원인)
            if data_split[0].strip().startswith('현상'):
                err_datalist.append(copy.deepcopy(cur_data))
                cur_data.reset()
                cur_col = data_split[0].strip()
                cur_data.set_data(col=cur_col, data=data_split[1].strip() + '\n')
            #                 errcode = extract_errcode(data_split[1])
            #                 if errcode in errcode_map:
            #                     print(errcode, errcode_map[errcode])
            else:
                cur_col = data_split[0].strip()
                cur_data.set_data(col=cur_col, data=data_split[1].strip() + '\n')
        else:
            pass  # 없음

for idx, err in enumerate(err_datalist):
    errcode_list = extract_errcode_list(err.situation, errcode_map)
    #     print(errcode_list)
    err.set_errcode(errcode_list)

# for err in err_datalist:
#     err.print_data()

from openpyxl import Workbook

write_wb = Workbook()

# oracle data 시트 생성
orcale_data_ws = write_wb.create_sheet('oracle_data')

# 행 단위로 추가
orcale_data_ws.append(['에러코드', '한글설명', '영어설명', '원인', '조치'])
for err in err_datalist:
    orcale_data_ws.append([err.error_code, err.description, err.situation, err.cause, err.solution])

# orcale error code map data 시트 생성
orcale_errcode_map_ws = write_wb.create_sheet('oracle_errcode_map')

# 행 단위로 추가
for code, desc in errcode_map.items():
    orcale_errcode_map_ws.append([code, desc])

# 저장
write_wb.save('dataset.xlsx')

